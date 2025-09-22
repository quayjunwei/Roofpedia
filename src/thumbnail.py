import os
import math
import requests
from io import BytesIO
from PIL import Image, ImageDraw
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from qgis.core import (
    QgsProject,
    QgsCoordinateReferenceSystem,
    QgsCoordinateTransform
)

# === CONFIGURATION ===
layer_name = "kanto_10000_centroid"
zoom = 18
tile_size = 256
crop_size = 512  # final image size (pixels)
grid_size = 7    # larger grid to ensure enough margin
stitched_size = tile_size * grid_size
tile_url_template = "https://mt1.google.com/vt/lyrs=y&x={x}&y={y}&z={z}"

output_base = "C:/Users/junwei.quay/Documents/qgis_japan/kanto_buildings/kanto_10000/img"
excel_path = "C:/Users/junwei.quay/Documents/qgis_japan/kanto_buildings/kanto_10000/kanto_roof_larger_than_10000_meters2.xlsx"
output_excel = "C:/Users/junwei.quay/Documents/qgis_japan/kanto_buildings/kanto_10000/kanto_roof_larger_than_10000_meters2_tn.xlsx"
thumbnail_column_letter = "I"
id_column_index = 1

os.makedirs(output_base, exist_ok=True)

# === HELPER FUNCTIONS ===
def latlon_to_tile_coords(lat, lon, zoom):
    lat_rad = math.radians(lat)
    n = 2.0 ** zoom
    x = (lon + 180.0) / 360.0 * n
    y = (1.0 - math.log(math.tan(lat_rad) + 1.0 / math.cos(lat_rad)) / math.pi) / 2.0 * n
    return x, y

# === LOAD QGIS LAYER ===
layer = QgsProject.instance().mapLayersByName(layer_name)[0]
crs_src = layer.crs()
crs_dest = QgsCoordinateReferenceSystem("EPSG:4326")
xform = QgsCoordinateTransform(crs_src, crs_dest, QgsProject.instance())

# === LOAD EXCEL ===
wb = load_workbook(excel_path)
ws = wb.active

# === BUILD FEATURE LOOKUP BY osm_id ===
feature_dict = {}
for feature in layer.getFeatures():
    osm_id = str(feature["osm_id"])
    feature_dict[osm_id] = feature

# === LOOP THROUGH EXCEL ROWS ===
headers = {"User-Agent": "Mozilla/5.0"}

for row_idx in range(2, ws.max_row + 1):
    osm_id = str(ws.cell(row=row_idx, column=id_column_index).value)
    if osm_id not in feature_dict:
        print(f"⚠️ No matching feature for osm_id {osm_id}")
        continue

    feature = feature_dict[osm_id]
    geom = feature.geometry().centroid().asPoint()
    wgs_point = xform.transform(geom)
    lat, lon = wgs_point.y(), wgs_point.x()

    # Get exact tile coordinates
    x_tile_float, y_tile_float = latlon_to_tile_coords(lat, lon, zoom)

    # Calculate top-left tile of stitched grid
    half_grid = grid_size // 2
    top_left_x = int(x_tile_float) - half_grid
    top_left_y = int(y_tile_float) - half_grid

    stitched_img = Image.new("RGB", (stitched_size, stitched_size))
    success = True

    for dx in range(grid_size):
        for dy in range(grid_size):
            tile_x = top_left_x + dx
            tile_y = top_left_y + dy
            tile_url = tile_url_template.format(x=tile_x, y=tile_y, z=zoom)
            try:
                response = requests.get(tile_url, headers=headers)
                if response.status_code == 200:
                    tile_img = Image.open(BytesIO(response.content)).convert("RGB")
                    stitched_img.paste(tile_img, (dx * tile_size, dy * tile_size))
                else:
                    print(f"❌ Failed to download tile {tile_x}, {tile_y}")
                    success = False
            except Exception as e:
                print(f"❌ Error downloading tile {tile_x}, {tile_y}: {e}")
                success = False

    if not success:
        ws.cell(row=row_idx, column=8).value = "Tile download failed"
        continue

    # Calculate pixel offset of centroid within stitched image
    pixel_x = int((x_tile_float - top_left_x) * tile_size)
    pixel_y = int((y_tile_float - top_left_y) * tile_size)

    # Crop fixed-size image centered on centroid
    left = pixel_x - crop_size // 2
    upper = pixel_y - crop_size // 2
    right = pixel_x + crop_size // 2
    lower = pixel_y + crop_size // 2
    cropped_img = stitched_img.crop((left, upper, right, lower))

    # Draw red cross at center
    draw = ImageDraw.Draw(cropped_img)
    cross_size = int(zoom * 1.5)
    center_x = crop_size // 2
    center_y = crop_size // 2
    draw.line((center_x - cross_size, center_y, center_x + cross_size, center_y), fill="red", width=3)
    draw.line((center_x, center_y - cross_size, center_x, center_y + cross_size), fill="red", width=3)

    # Save image
    output_path = os.path.join(output_base, f"{osm_id}.jpeg")
    cropped_img.save(output_path)
    print(f"✅ Saved centered image for {osm_id}")

    # Insert into Excel
    try:
        excel_img = ExcelImage(output_path)
        excel_img.width = crop_size
        excel_img.height = crop_size
        ws.add_image(excel_img, f"{thumbnail_column_letter}{row_idx}")
        print(f"✅ Inserted thumbnail for {osm_id}")

        # Resize row and column to fit image
        row_height = crop_size / 1.33
        column_width = crop_size / 7
        ws.row_dimensions[row_idx].height = row_height
        ws.column_dimensions[thumbnail_column_letter].width = column_width

    except Exception as e:
        print(f"❌ Failed to insert image into Excel for {osm_id}: {e}")
        ws.cell(row=row_idx, column=8).value = "Image not found"

# === SAVE EXCEL ===
wb.save(output_excel)
print("🎯 All images inserted, cells resized, and Excel saved.")